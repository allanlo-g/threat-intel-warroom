# -*- coding: utf-8 -*-
"""
Snapshot generator for the Threat-Intel War-Room dashboard.

Reads config/companies.json, downloads each company's yearly Google Sheets
(xlsx export), parses the monthly tabs and writes a normalized JSON snapshot
to data/snapshot.<id>.json. The dashboard reads Google Sheets live in the
browser; this snapshot is the offline / GitHub-Pages fallback.

Usage:
    python scripts/snapshot.py                # all active companies
    python scripts/snapshot.py --id demo      # one company
    python scripts/snapshot.py --use-local    # parse already-downloaded _y*.xlsx (dev)

Requires: openpyxl  (pip install openpyxl)
"""
import argparse
import io
import json
import os
import re
import sys
import urllib.request

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MONTH_RE = re.compile(r"^\s*(\d{4})\s*年\s*(\d{1,2})\s*月\s*$")

# canonical column order of a monthly tab (11 columns)
FIELDS = [
    "date", "source", "title", "url", "origin", "category",
    "severity", "riskScope", "affected", "countermeasure", "progress",
]


def log(msg):
    sys.stdout.write(msg + "\n")
    sys.stdout.flush()


def export_url(sheet_id):
    return "https://docs.google.com/spreadsheets/d/%s/export?format=xlsx" % sheet_id


def download_xlsx(sheet_id, timeout=60):
    req = urllib.request.Request(export_url(sheet_id), headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read()


def cell_str(v):
    if v is None:
        return ""
    # openpyxl may hand back datetime for date cells
    if hasattr(v, "strftime"):
        try:
            return v.strftime("%Y/%-m/%-d")
        except ValueError:  # Windows strftime has no %-m
            return "%d/%d/%d" % (v.year, v.month, v.day)
    return str(v).strip()


def parse_workbook(xlsx_bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    records = []
    for sn in wb.sheetnames:
        m = MONTH_RE.match(sn)
        if not m:
            continue  # skip 說明 / 資安情資來源 / 外部議題蒐集 ... helper tabs
        year, month = int(m.group(1)), int(m.group(2))
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        for r in rows[1:]:  # skip header
            cells = [cell_str(c) for c in r]
            if not any(cells):
                continue
            # title is the strongest signal that a row is real data
            title = cells[2] if len(cells) > 2 else ""
            if not title:
                continue
            rec = {}
            for i, f in enumerate(FIELDS):
                rec[f] = cells[i] if i < len(cells) else ""
            rec["year"] = year
            rec["month"] = month
            # derive day from date when possible
            day = ""
            dm = re.search(r"(\d{1,4})[/-](\d{1,2})[/-](\d{1,2})", rec["date"])
            if dm:
                day = int(dm.group(3))
            rec["day"] = day
            records.append(rec)
    return records


def build_snapshot(company, use_local):
    cid = company["id"]
    years = company.get("years", {})
    by_year = {}
    for y, sid in sorted(years.items()):
        if use_local:
            path = os.path.join(ROOT, "_y%s.xlsx" % y)
            if not os.path.exists(path):
                log("  [%s] %s: no local file, skip" % (cid, y))
                continue
            data = open(path, "rb").read()
        else:
            if not sid:
                log("  [%s] %s: no sheet id, skip" % (cid, y))
                continue
            try:
                log("  [%s] %s: downloading ..." % (cid, y))
                data = download_xlsx(sid)
            except Exception as e:  # noqa
                log("  [%s] %s: download failed (%s)" % (cid, y, e))
                continue
        recs = parse_workbook(data)
        by_year[str(y)] = recs
        log("  [%s] %s: %d records" % (cid, y, len(recs)))
    return {
        "company": cid,
        "name": company.get("name", cid),
        "years": by_year,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--id", help="only this company id")
    ap.add_argument("--use-local", action="store_true",
                    help="parse already-downloaded _y<year>.xlsx instead of Google")
    args = ap.parse_args()

    cfg_path = os.path.join(ROOT, "config", "companies.json")
    cfg = json.load(io.open(cfg_path, encoding="utf-8"))
    companies = cfg.get("companies", [])
    if args.id:
        companies = [c for c in companies if c["id"] == args.id]
    if not companies:
        log("no matching company")
        return

    for c in companies:
        if not c.get("active", True) and not args.id:
            continue
        log("== %s (%s) ==" % (c.get("name", c["id"]), c["id"]))
        snap = build_snapshot(c, args.use_local)
        out = os.path.join(ROOT, "data", "snapshot.%s.json" % c["id"])
        with io.open(out, "w", encoding="utf-8") as f:
            json.dump(snap, f, ensure_ascii=False, separators=(",", ":"))
        total = sum(len(v) for v in snap["years"].values())
        log("  -> %s  (%d records total)" % (os.path.relpath(out, ROOT), total))


if __name__ == "__main__":
    main()
